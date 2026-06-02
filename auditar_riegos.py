"""
auditar_riegos.py — Auditoría completa de riego, temporada 2025-2026.
Uso: python auditar_riegos.py [equipo]
      python auditar_riegos.py 2
      python auditar_riegos.py          # usa EQUIPO por defecto (1)
"""

import sqlite3, openpyxl, sys
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "riego.db"
DESGLOSE = BASE_DIR / "Cuartel x Sector.xlsx"
HISTORIAL = BASE_DIR / "Historial"
EQUIPO = int(sys.argv[1]) if len(sys.argv) > 1 else 1
OUTPUT = BASE_DIR / f"Auditoria_Riego_Equipo{EQUIPO}.xlsx"

# =============================================
# 1. BD
# =============================================
conn = sqlite3.connect(str(DB_PATH))
conn.execute("PRAGMA foreign_keys = ON")
conn.executescript("""
DROP TABLE IF EXISTS riegos_cuartel;
DROP TABLE IF EXISTS riegos;
DROP TABLE IF EXISTS cuartel_sector;
DROP TABLE IF EXISTS cuarteles;
DROP TABLE IF EXISTS sectores;

CREATE TABLE cuarteles (
    cc INTEGER PRIMARY KEY, variedad TEXT NOT NULL, anio INTEGER,
    dist_hilera REAL, dist_plantas REAL, has_total REAL
);
CREATE TABLE sectores (
    sector_nom TEXT PRIMARY KEY, equipo INTEGER NOT NULL, sector_num INTEGER NOT NULL,
    caudal_nominal REAL, has_total REAL, caseta TEXT
);
CREATE TABLE cuartel_sector (
    cuartel_id INTEGER NOT NULL REFERENCES cuarteles(cc),
    sector_nom TEXT NOT NULL REFERENCES sectores(sector_nom),
    equipo INTEGER NOT NULL, sector_num INTEGER NOT NULL,
    porcentaje REAL NOT NULL, has_en_sector REAL NOT NULL,
    PRIMARY KEY (cuartel_id, sector_nom)
);
CREATE TABLE riegos (
    id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL,
    equipo_num INTEGER NOT NULL, sector_nom TEXT NOT NULL,
    volumen_m3 REAL NOT NULL, duracion_h REAL,
    hora_ini TEXT, hora_fin TEXT, archivo_origen TEXT
);
CREATE TABLE riegos_cuartel (
    riego_id INTEGER NOT NULL REFERENCES riegos(id),
    cuartel_id INTEGER NOT NULL REFERENCES cuarteles(cc),
    sector_nom TEXT NOT NULL, volumen_m3 REAL NOT NULL
);
CREATE INDEX idx_riegos_sec ON riegos(sector_nom);
CREATE INDEX idx_riegos_fecha ON riegos(fecha);
CREATE INDEX idx_rc_cuar ON riegos_cuartel(cuartel_id);
CREATE INDEX idx_rc_riego ON riegos_cuartel(riego_id);
""")
print("[1/5] BD creada")

# =============================================
# 2. Importar Sectores y Cuartel x Sector
# =============================================
wb = openpyxl.load_workbook(str(DESGLOSE), data_only=True)

ws_sec = wb["Sectores"]
for row in ws_sec.iter_rows(min_row=2, values_only=True):
    eq, snum, caudal, has_tot, caseta = row[0], row[1], row[2], row[3], row[4]
    if eq is None:
        continue
    eq = int(eq)
    if eq != EQUIPO:
        continue
    snum = int(snum)
    sn = f"E{eq}S{snum}"
    conn.execute(
        "INSERT INTO sectores VALUES (?,?,?,?,?,?)",
        (sn, eq, snum, caudal, has_tot, str(caseta) if caseta else None),
    )

ws_cs = wb["Cuartel x Sector"]
for row in ws_cs.iter_rows(min_row=2, values_only=True):
    cc, var, dh, dp, anio, eq, snum, pct, has_cuar, has_sec = row
    if cc is None or eq is None:
        continue
    eq = int(eq)
    if eq != EQUIPO:
        continue
    cc = int(cc)
    snum = int(snum)
    sn = f"E{eq}S{snum}"
    # Auto-crear sector si no existe
    if not conn.execute(
        "SELECT 1 FROM sectores WHERE sector_nom = ?", (sn,)
    ).fetchone():
        conn.execute(
            "INSERT OR IGNORE INTO sectores VALUES (?,?,?,?,?,?)",
            (sn, eq, snum, None, 0.0, ""),
        )
    conn.execute(
        "INSERT OR REPLACE INTO cuarteles VALUES (?,?,?,?,?,?)",
        (cc, str(var), int(anio) if anio else None, dh, dp, has_cuar),
    )
    conn.execute(
        "INSERT OR REPLACE INTO cuartel_sector VALUES (?,?,?,?,?,?)",
        (cc, sn, eq, snum, pct, has_sec),
    )

conn.commit()
n1 = conn.execute("SELECT COUNT(*) FROM cuarteles").fetchone()[0]
n2 = conn.execute("SELECT COUNT(*) FROM cuartel_sector").fetchone()[0]
print(f"[2/5] Importados: {n1} cuarteles, {n2} relaciones")

# =============================================
# 3. Historial de riegos
# =============================================
importados = 0
for arch in sorted(HISTORIAL.glob("*.xlsx")):
    if arch.name.startswith("~$"):
        continue
    wh = openpyxl.load_workbook(str(arch), data_only=True)
    if "Riegos" not in wh.sheetnames:
        continue
    ws_r = wh["Riegos"]
    hdrs = [c.value for c in ws_r[1]]
    es_c = "Tipo" in hdrs

    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if es_c:
            f, eqn, sn = row[0], row[2], row[3]
            hi, hf, d, v = row[4], row[5], row[6], row[7]
            if eqn is None or sn is None:
                continue
            eqn = int(eqn)
            sn = f"E{eqn}S{int(sn)}"
        else:
            f, eqn, sn = row[0], row[1], row[4]
            hi, hf, d, v = row[5], row[6], row[7], row[8]
            if eqn is None or sn is None:
                continue
            eqn = int(eqn)

        if eqn != EQUIPO:
            continue
        conn.execute(
            "INSERT INTO riegos VALUES (NULL,?,?,?,?,?,?,?,?)",
            (
                str(f) if f else None,
                eqn,
                str(sn).strip(),
                float(v) if v else 0,
                float(d) if d else None,
                str(hi) if hi else None,
                str(hf) if hf else None,
                arch.name,
            ),
        )
        importados += 1

conn.commit()
print(f"[3/5] Historial: {importados} riegos")

# =============================================
# 4. Distribuir m3 por cuartel
# =============================================
conn.row_factory = sqlite3.Row

# Pre-cargar cuarteles por sector
cporsec = {}
for cs in conn.execute("SELECT * FROM cuartel_sector").fetchall():
    s = cs["sector_nom"]
    if s not in cporsec:
        cporsec[s] = []
    cporsec[s].append({"cc": cs["cuartel_id"], "ha": cs["has_en_sector"]})

sum_ha = {s: sum(c["ha"] for c in lista) for s, lista in cporsec.items()}

dist = 0
for riego in conn.execute("SELECT * FROM riegos").fetchall():
    s = riego["sector_nom"]
    if s not in cporsec:
        continue
    sh = sum_ha[s]
    if sh == 0:
        continue
    for c in cporsec[s]:
        m3 = riego["volumen_m3"] * (c["ha"] / sh)
        conn.execute(
            "INSERT INTO riegos_cuartel VALUES (?,?,?,?)",
            (riego["id"], c["cc"], s, round(m3, 4)),
        )
    dist += 1

conn.commit()
n3 = conn.execute("SELECT COUNT(*) FROM riegos_cuartel").fetchone()[0]
print(f"[4/5] Distribucion: {dist} riegos -> {n3} registros cuartel")

# =============================================
# 5. Exportar Excel
# =============================================
wb_out = openpyxl.Workbook()
H_FONT = Font(bold=True, color="FFFFFF", size=11)
H_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
B_ALIGN = Alignment(vertical="center")
BDR = Border(
    left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
)
ALT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def sh(ws, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = H_FONT, H_FILL, H_ALIGN, BDR


def aw(ws, nc, nr):
    for c in range(1, nc + 1):
        mx = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, nr + 1)),
            default=0,
        )
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, 32)


# --- Hoja 1: Sectores ---
ws1 = wb_out.active
ws1.title = "Sectores"
h1 = ["Sector", "Equipo", "Nro Sector", "Ha Totales", "Caudal (m3/h)", "Caseta"]
for i, h in enumerate(h1, 1):
    ws1.cell(row=1, column=i, value=h)
rows1 = conn.execute("SELECT * FROM sectores ORDER BY sector_nom").fetchall()
for r, row in enumerate(rows1, 2):
    for c, k in enumerate(
        ["sector_nom", "equipo", "sector_num", "has_total", "caudal_nominal", "caseta"],
        1,
    ):
        cell = ws1.cell(row=r, column=c, value=row[k])
        cell.border = BDR
        if k == "has_total":
            cell.number_format = "0.00"
sh(ws1, 6)
aw(ws1, 6, len(rows1) + 1)
ws1.auto_filter.ref = ws1.dimensions
ws1.freeze_panes = "A2"

# --- Hoja 2: Cuartel x Sector ---
ws2 = wb_out.create_sheet("Cuartel_x_Sector")
h2 = [
    "Cuartel (CC)",
    "Variedad",
    "Anio",
    "Dist. Hilera",
    "Dist. Plantas",
    "Ha Cuartel",
    "Sector",
    "% del Cuartel",
    "Ha en Sector",
]
for i, h in enumerate(h2, 1):
    ws2.cell(row=1, column=i, value=h)
rows2 = conn.execute("""
    SELECT cs.cuartel_id, c.variedad, c.anio, c.dist_hilera, c.dist_plantas,
           c.has_total, cs.sector_nom, cs.porcentaje, cs.has_en_sector
    FROM cuartel_sector cs JOIN cuarteles c ON cs.cuartel_id = c.cc
    ORDER BY cs.cuartel_id, cs.sector_nom
""").fetchall()
for r, row in enumerate(rows2, 2):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.border = BDR
        cell.alignment = B_ALIGN
        if c in (6, 9):
            cell.number_format = "0.00"
        if c == 8:
            cell.number_format = "0%"
    if r % 2 == 0:
        for c in range(1, 10):
            ws2.cell(row=r, column=c).fill = ALT
sh(ws2, 9)
aw(ws2, 9, len(rows2) + 1)
ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = "A2"

# --- Hoja 3: Historial Riegos ---
ws3 = wb_out.create_sheet("Historial_Riegos")
h3 = [
    "ID",
    "Fecha",
    "Sector",
    "Volumen (m3)",
    "Duracion (h)",
    "Hora Ini",
    "Hora Fin",
    "Archivo",
]
for i, h in enumerate(h3, 1):
    ws3.cell(row=1, column=i, value=h)
rows3 = conn.execute("""
    SELECT id, fecha, sector_nom, volumen_m3, duracion_h, hora_ini, hora_fin, archivo_origen
    FROM riegos ORDER BY fecha
""").fetchall()
for r, row in enumerate(rows3, 2):
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.border = BDR
        if c in (4, 5):
            cell.number_format = "#,##0.0"
    if r % 2 == 0:
        for c in range(1, 9):
            ws3.cell(row=r, column=c).fill = ALT
sh(ws3, 8)
aw(ws3, 8, len(rows3) + 1)
ws3.auto_filter.ref = ws3.dimensions
ws3.freeze_panes = "A2"

# --- Hoja 4: Resumen Cuartel x Mes ---
ws4 = wb_out.create_sheet("Resumen_Cuartel_x_Mes")

meses = conn.execute(
    "SELECT DISTINCT strftime('%Y-%m', fecha) as mes FROM riegos ORDER BY mes"
).fetchall()
mes_raw = [m["mes"] for m in meses]
M_ES = {
    "01": "Enero",
    "02": "Febrero",
    "03": "Marzo",
    "04": "Abril",
    "05": "Mayo",
    "06": "Junio",
    "07": "Julio",
    "08": "Agosto",
    "09": "Septiembre",
    "10": "Octubre",
    "11": "Noviembre",
    "12": "Diciembre",
}
mes_disp = []
for ml in mes_raw:
    a, n = ml.split("-")
    mes_disp.append(f"{M_ES[n]} {a}")

cuarteles = conn.execute("""
    SELECT rc.cuartel_id, c.variedad, c.has_total,
           (SELECT SUM(cs.has_en_sector) FROM cuartel_sector cs WHERE cs.cuartel_id = rc.cuartel_id) as ha_regada,
           GROUP_CONCAT(DISTINCT rc.sector_nom ORDER BY rc.sector_nom) as sectores
    FROM riegos_cuartel rc JOIN cuarteles c ON rc.cuartel_id = c.cc
    GROUP BY rc.cuartel_id ORDER BY rc.cuartel_id
""").fetchall()

m3_dict = {}
for row in conn.execute("""
    SELECT rc.cuartel_id, strftime('%Y-%m', r.fecha) as mes, ROUND(SUM(rc.volumen_m3), 1) as m3
    FROM riegos_cuartel rc JOIN riegos r ON rc.riego_id = r.id
    GROUP BY rc.cuartel_id, mes
""").fetchall():
    m3_dict[(row["cuartel_id"], row["mes"])] = row["m3"]

MI = 5  # columna donde empiezan los meses
h4 = (
    ["Cuartel", "Variedad", "Ha Regada", "Sector(es)"]
    + [f"{m} (m3)" for m in mes_disp]
    + ["Total (m3)", "m3/ha"]
)
nc4 = len(h4)
for i, h in enumerate(h4, 1):
    ws4.cell(row=1, column=i, value=h)

for r, cd in enumerate(cuarteles, 2):
    cc = cd["cuartel_id"]
    has = cd["ha_regada"]
    sn = cd["sectores"]
    for c, v in enumerate([cc, cd["variedad"], has, sn], 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.border = BDR
        if c == 3:
            cell.number_format = "0.00"
    total_cc = 0
    for ci, mk in enumerate(mes_raw):
        m3 = m3_dict.get((cc, mk), 0)
        cell = ws4.cell(row=r, column=MI + ci, value=m3 if m3 else None)
        cell.border = BDR
        cell.alignment = B_ALIGN
        cell.number_format = "#,##0.0"
        if m3:
            total_cc += m3
    ct = ws4.cell(row=r, column=MI + len(mes_raw), value=round(total_cc, 1))
    ct.border = BDR
    ct.font = Font(bold=True)
    ct.number_format = "#,##0.0"
    ch = ws4.cell(
        row=r,
        column=MI + len(mes_raw) + 1,
        value=round(total_cc / has, 1) if has else None,
    )
    ch.border = BDR
    ch.font = Font(bold=True)
    ch.number_format = "#,##0.0"
    if r % 2 == 0:
        for c in range(1, nc4 + 1):
            ws4.cell(row=r, column=c).fill = ALT

# Fila totales
tr = len(cuarteles) + 2
for c in range(1, nc4 + 1):
    ws4.cell(row=tr, column=c).fill = H_FILL
    ws4.cell(row=tr, column=c).border = BDR
    ws4.cell(row=tr, column=c).font = Font(bold=True, color="FFFFFF")
ws4.cell(row=tr, column=1, value="TOTAL")
for ci, mk in enumerate(mes_raw):
    s = sum(m3_dict.get((cd["cuartel_id"], mk), 0) for cd in cuarteles)
    cell = ws4.cell(row=tr, column=MI + ci, value=round(s, 1))
    cell.number_format = "#,##0.0"

sh(ws4, nc4)
aw(ws4, nc4, tr + 1)
ws4.auto_filter.ref = f"A1:{get_column_letter(nc4)}{tr - 1}"
ws4.freeze_panes = "E2"

# --- Hoja 5: Resumen Detallado Cuartel x Mes (m3 + m3/ha por mes) ---
ws5 = wb_out.create_sheet("Resumen_Detallado")

meses_inicio_5 = 5
h5 = ["Cuartel", "Variedad", "Ha Regada", "Sector(es)"]
for m in mes_disp:
    h5.append(f"{m} (m3)")
    h5.append(f"{m} (m3/ha)")
h5 += ["Total (m3)", "m3/ha temp"]
nc5 = len(h5)
for i, h in enumerate(h5, 1):
    ws5.cell(row=1, column=i, value=h)

for r, cd in enumerate(cuarteles, 2):
    cc = cd["cuartel_id"]
    has = cd["ha_regada"]
    sn = cd["sectores"]
    for c, v in enumerate([cc, cd["variedad"], has, sn], 1):
        cell = ws5.cell(row=r, column=c, value=v)
        cell.border = BDR
        if c == 3:
            cell.number_format = "0.00"
    total_cc = 0
    for ci, mk in enumerate(mes_raw):
        col_m3 = meses_inicio_5 + ci * 2
        col_mha = meses_inicio_5 + ci * 2 + 1
        m3 = m3_dict.get((cc, mk), 0)
        c_m3 = ws5.cell(row=r, column=col_m3, value=m3 if m3 else None)
        c_m3.border = BDR
        c_m3.alignment = B_ALIGN
        c_m3.number_format = "#,##0.0"
        m3ha_mes = round(m3 / has, 1) if has and m3 else None
        c_mha = ws5.cell(row=r, column=col_mha, value=m3ha_mes)
        c_mha.border = BDR
        c_mha.alignment = B_ALIGN
        c_mha.number_format = "#,##0.0"
        c_mha.font = Font(bold=True, size=12)
        if m3:
            total_cc += m3
    ct = ws5.cell(
        row=r, column=meses_inicio_5 + len(mes_raw) * 2, value=round(total_cc, 1)
    )
    ct.border = BDR
    ct.font = Font(bold=True)
    ct.number_format = "#,##0.0"
    ch = ws5.cell(
        row=r,
        column=meses_inicio_5 + len(mes_raw) * 2 + 1,
        value=round(total_cc / has, 1) if has else None,
    )
    ch.border = BDR
    ch.font = Font(bold=True)
    ch.number_format = "#,##0.0"
    if r % 2 == 0:
        for c in range(1, nc5 + 1):
            ws5.cell(row=r, column=c).fill = ALT

# Fila totales hoja 5
tr5 = len(cuarteles) + 2
for c in range(1, nc5 + 1):
    ws5.cell(row=tr5, column=c).fill = H_FILL
    ws5.cell(row=tr5, column=c).border = BDR
    ws5.cell(row=tr5, column=c).font = Font(bold=True, color="FFFFFF")
ws5.cell(row=tr5, column=1, value="TOTAL")
for ci, mk in enumerate(mes_raw):
    col_m3 = meses_inicio_5 + ci * 2
    s = sum(m3_dict.get((cd["cuartel_id"], mk), 0) for cd in cuarteles)
    cell = ws5.cell(row=tr5, column=col_m3, value=round(s, 1))
    cell.number_format = "#,##0.0"

sh(ws5, nc5)
aw(ws5, nc5, tr5 + 1)
ws5.auto_filter.ref = f"A1:{get_column_letter(nc5)}{tr5 - 1}"
ws5.freeze_panes = "E2"

wb_out.save(str(OUTPUT))
print(f"[5/5] Excel exportado: {OUTPUT.name}")
print(f"     Sectores: {len(rows1)}")
print(f"     Cuartel x Sector: {len(rows2)}")
print(f"     Historial: {len(rows3)}")
print(f"     Resumen: {len(cuarteles)} cuarteles x {len(mes_raw)} meses")

conn.close()
