import streamlit as st, pandas as pd, sqlite3, openpyxl, re, json
from io import BytesIO
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Riego por Cuartel Siracusa", layout="wide")
st.title("Riego por Cuartel Siracusa")

# ─── CONFIG ───────────────────────────────────────────
LATCH_EQUIPOS = [10, 11, 19]
LATCH_SECTOR_MAP = {
    10: {
        "1": 1,
        "2": 1,
        "3": 1,
        "17": 2,
        "7": 3,
        "8": 3,
        "9": 3,
        "10": 4,
        "11": 4,
        "12": 4,
        "15": 5,
        "18": 6,
    },
    11: {
        "7": 1,
        "6": 1,
        "10": 1,
        "9": 1,
        "5": 2,
        "12": 2,
        "13": 2,
        "17": 2,
        "11": 3,
        "14": 3,
        "15": 3,
        "8": 4,
        "16": 4,
        "1": 5,
        "2": 5,
        "3": 5,
        "4": 5,
    },
    19: {
        "1": 1,
        "2": 2,
        "3": 3,
        "4": 3,
        "5": 3,
        "6": 4,
        "7": 4,
        "8": 4,
        "9": 4,
        "10": 5,
        "11": 5,
        "12": 5,
        "13": 5,
        "14": 6,
    },
}
IMPULSION_SECTORES = {1: [6, 7, 8], 2: [6], 3: [8], 9: [9], 10: [15, 16]}
MESES_ES = {
    "01": "Enero",
    "02": "Febrero",
    "03": "Marzo",
    "04": "Abril",
    "05": "Mayo",
    "06": "Junio",
    "07": "Julio",
    "08": "Agosto",
    "09": "Septiembre",
    "10": "Octubre",
    "11": "Noviembre",
    "12": "Diciembre",
}

# ─── CARGA INICIAL DE DATOS ───────────────────────────
DATA_JSON = Path(__file__).parent / "data_base.json"


def load_sectores():
    if "sectores_df" not in st.session_state:
        with open(DATA_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        st.session_state.sectores_df = pd.DataFrame(data["sectores"])
        st.session_state.cuartel_sector_df = pd.DataFrame(data["cuartel_sector"])
    return st.session_state.sectores_df, st.session_state.cuartel_sector_df


def save_session_data():
    sectores_df, cuartel_sector_df = load_sectores()
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sectores_df.to_excel(writer, sheet_name="Sectores", index=False)
        cs_df = cuartel_sector_df.copy()
        cs_df.columns = [
            "ID",
            "CC",
            "Variedad",
            "Dist Hilera",
            "Dist Plantas",
            "Anio",
            "Equipo",
            "Sector",
            "%",
            "Ha Cuartel",
            "Ha en Sector",
            "Activo",
        ]
        cs_df.to_excel(writer, sheet_name="Cuartel x Sector", index=False)
    buf.seek(0)
    return buf


# ─── CLEANER ──────────────────────────────────────────
def parse_duration_hours(value):
    if pd.isna(value):
        return None
    parts = str(value).strip().split(":")
    try:
        if len(parts) == 3:
            return int(parts[0]) + int(parts[1]) / 60 + int(parts[2]) / 3600
        if len(parts) == 2:
            return int(parts[0]) / 60 + int(parts[1]) / 3600
        if len(parts) == 1:
            return float(parts[0])
    except:
        return None
    return None


def extraer_numero_equipo(equipo_str):
    if pd.isna(equipo_str):
        return None
    m = re.search(r"E(\d+)$", str(equipo_str).strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


def clean_agronic(uploaded_file):
    df = pd.read_excel(uploaded_file, engine="calamine")
    df["fecha"] = pd.to_datetime(df["Desde"], dayfirst=True, errors="coerce")
    df["hora_inicio"] = df["fecha"].dt.time
    df["fecha_fin"] = pd.to_datetime(df["Hasta"], dayfirst=True, errors="coerce")
    df["hora_fin"] = df["fecha_fin"].dt.time
    df["horas"] = df["Tiempo"].apply(parse_duration_hours)
    df["equipo_num"] = df["Equipo"].apply(extraer_numero_equipo)

    def get_sector_real(eq_num, subsector):
        if eq_num in LATCH_SECTOR_MAP:
            s = str(int(subsector)) if subsector is not None else None
            if s in LATCH_SECTOR_MAP[eq_num]:
                return LATCH_SECTOR_MAP[eq_num][s]
        return None

    def es_impulsion(eq_num, sector):
        return eq_num in IMPULSION_SECTORES and sector in IMPULSION_SECTORES[eq_num]

    def clasificar(row):
        eq = row["equipo_num"]
        sec = row["Sector"]
        eq_str = str(row.get("Equipo", ""))
        if eq_str.strip() == "POZO":
            return "IMPULSION"
        if eq in LATCH_EQUIPOS:
            return "IMPULSION" if es_impulsion(eq, sec) else "RIEGO_LATCH"
        return "IMPULSION" if es_impulsion(eq, sec) else "RIEGO"

    df["tipo"] = df.apply(clasificar, axis=1)

    def mapear_sector(row):
        eq = row["equipo_num"]
        sec = row["Sector"]
        if eq in LATCH_EQUIPOS:
            return get_sector_real(eq, sec)
        return sec

    df["sector_mapeado"] = df.apply(mapear_sector, axis=1)

    df_riego = df[df["tipo"].isin(["RIEGO", "RIEGO_LATCH"])].copy()
    df_latch = df_riego[df_riego["tipo"] == "RIEGO_LATCH"]
    df_normal = df_riego[df_riego["tipo"] == "RIEGO"]

    if len(df_latch) > 0:
        latch = (
            df_latch.groupby(["fecha", "equipo_num", "sector_mapeado"])
            .agg(
                {
                    "hora_inicio": "min",
                    "hora_fin": "max",
                    "horas": "max",
                    "Volumen (m3)": "sum",
                }
            )
            .reset_index()
        )
        latch["tipo"] = "RIEGO_LATCH"
    else:
        latch = pd.DataFrame()

    riego_final = (
        pd.concat([df_normal, latch], ignore_index=True)
        if len(latch) > 0
        else df_normal.copy()
    )
    planilla = []
    for _, row in riego_final.iterrows():
        eqn = row.get("equipo_num")
        sec = row.get("sector_mapeado")
        vol = row.get("Volumen (m3)", 0) or 0
        equipo_str = (
            str(row.get("Equipo", "")).split()[0] if pd.notna(row.get("Equipo")) else ""
        )
        planilla.append(
            {
                "Fecha": row["fecha"],
                "Equipo": equipo_str,
                "Equipo_Num": int(eqn) if pd.notna(eqn) else None,
                "Sector": int(sec) if pd.notna(sec) else None,
                "Hora_Ini": row["hora_inicio"],
                "Hora_Fin": row["hora_fin"],
                "Duracion_h": round(row["horas"], 2)
                if pd.notna(row.get("horas"))
                else 0,
                "Volumen_m3": round(float(vol), 2),
                "Tipo": row.get("tipo", "RIEGO"),
            }
        )
    return pd.DataFrame(planilla)


# ─── AUDITOR ──────────────────────────────────────────
def run_audit(sectores_df, cuartel_sector_df, riego_files, equipo):
    conn = sqlite3.connect(":memory:")
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript("""
    CREATE TABLE cuarteles (cc INTEGER PRIMARY KEY, variedad TEXT NOT NULL, anio INTEGER, dist_hilera REAL, dist_plantas REAL, has_total REAL);
    CREATE TABLE sectores (sector_nom TEXT PRIMARY KEY, equipo INTEGER NOT NULL, sector_num INTEGER NOT NULL, caudal_nominal REAL, has_total REAL, caseta TEXT);
    CREATE TABLE cuartel_sector (cuartel_id INTEGER NOT NULL REFERENCES cuarteles(cc), sector_nom TEXT NOT NULL REFERENCES sectores(sector_nom), equipo INTEGER NOT NULL, sector_num INTEGER NOT NULL, porcentaje REAL NOT NULL, has_en_sector REAL NOT NULL, PRIMARY KEY (cuartel_id, sector_nom));
    CREATE TABLE riegos (id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL, equipo_num INTEGER NOT NULL, sector_nom TEXT NOT NULL, volumen_m3 REAL NOT NULL, duracion_h REAL, hora_ini TEXT, hora_fin TEXT, archivo_origen TEXT);
    CREATE TABLE riegos_cuartel (riego_id INTEGER NOT NULL REFERENCES riegos(id), cuartel_id INTEGER NOT NULL REFERENCES cuarteles(cc), sector_nom TEXT NOT NULL, volumen_m3 REAL NOT NULL);
    """)

    # Cargar sectores desde session state
    for _, row in sectores_df.iterrows():
        if equipo is not None and row["equipo"] != equipo:
            continue
        sn = f"E{int(row['equipo'])}S{int(row['sector'])}"
        conn.execute(
            "INSERT INTO sectores VALUES (?,?,?,?,?,?)",
            (
                sn,
                int(row["equipo"]),
                int(row["sector"]),
                row.get("caudal"),
                row["has"],
                str(row.get("caseta", "")),
            ),
        )

    # Cargar cuartel x sector desde session state (solo activos)
    for _, row in cuartel_sector_df.iterrows():
        if equipo is not None and row["equipo"] != equipo:
            continue
        if "activo" in row and not row["activo"]:
            continue
        cc, eq, snum = int(row["cc"]), int(row["equipo"]), int(row["sector"])
        sn = f"E{eq}S{snum}"
        # Auto-crear sector si no existe en la tabla sectores
        existing = conn.execute(
            "SELECT 1 FROM sectores WHERE sector_nom = ?", (sn,)
        ).fetchone()
        if not existing:
            conn.execute(
                "INSERT OR IGNORE INTO sectores VALUES (?,?,?,?,?,?)",
                (sn, eq, snum, None, 0.0, ""),
            )
        conn.execute(
            "INSERT OR REPLACE INTO cuarteles VALUES (?,?,?,?,?,?)",
            (
                cc,
                str(row["variedad"]),
                int(row["anio"]) if pd.notna(row["anio"]) else None,
                row.get("dh"),
                row.get("dp"),
                row["has_total"],
            ),
        )
        conn.execute(
            "INSERT OR REPLACE INTO cuartel_sector VALUES (?,?,?,?,?,?)",
            (cc, sn, eq, snum, row["pct"], row["has_en_sector"]),
        )

    # Cargar historial de riegos
    for rf in riego_files:
        wb_hist = openpyxl.load_workbook(BytesIO(rf.read()), data_only=True)
        if "Riegos" not in wb_hist.sheetnames:
            continue
        ws_r = wb_hist["Riegos"]
        hdrs = [c.value for c in ws_r[1]]
        es_c = hdrs and "Tipo" in str(hdrs)
        for row in ws_r.iter_rows(min_row=2, values_only=True):
            if es_c:
                f, eqn, sec_num = row[0], row[2], row[3]
                hi, hf, d, v = row[4], row[5], row[6], row[7]
                if eqn is None or sec_num is None:
                    continue
                eqn = int(eqn)
                sn = f"E{eqn}S{int(sec_num)}"
            else:
                f, eqn, sn = row[0], row[1], row[4]
                hi, hf, d, v = row[5], row[6], row[7], row[8]
                if eqn is None or sn is None:
                    continue
                eqn = int(eqn)
            if equipo is not None and eqn != equipo:
                continue
            conn.execute(
                "INSERT INTO riegos VALUES (NULL,?,?,?,?,?,?,?,?)",
                (
                    str(f) if f else None,
                    eqn,
                    str(sn).strip(),
                    float(v) if v else 0,
                    float(d) if d else None,
                    str(hi) if hi else None,
                    str(hf) if hf else None,
                    rf.name,
                ),
            )
    conn.commit()

    # Distribuir
    conn.row_factory = sqlite3.Row
    cs_by_sector = {}
    for cs in conn.execute("SELECT * FROM cuartel_sector").fetchall():
        s = cs["sector_nom"]
        if s not in cs_by_sector:
            cs_by_sector[s] = []
        cs_by_sector[s].append({"cc": cs["cuartel_id"], "ha": cs["has_en_sector"]})
    sum_ha = {s: sum(c["ha"] for c in lst) for s, lst in cs_by_sector.items()}

    for riego in conn.execute("SELECT * FROM riegos").fetchall():
        s = riego["sector_nom"]
        if s not in cs_by_sector or sum_ha.get(s, 0) == 0:
            continue
        for c in cs_by_sector[s]:
            m3 = riego["volumen_m3"] * (c["ha"] / sum_ha[s])
            conn.execute(
                "INSERT INTO riegos_cuartel VALUES (?,?,?,?)",
                (riego["id"], c["cc"], s, round(m3, 4)),
            )
    conn.commit()

    # Construir Excel output
    wb_out = openpyxl.Workbook()
    HF = Font(bold=True, color="FFFFFF", size=11)
    HFL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BA = Alignment(vertical="center")
    BDR = Border(
        left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
    )
    ALT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    def style_h(ws, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill, cell.alignment, cell.border = HF, HFL, HA, BDR

    def auto_w(ws, nc, nr):
        for c in range(1, nc + 1):
            mx = max(
                (
                    len(str(ws.cell(row=r, column=c).value or ""))
                    for r in range(1, nr + 1)
                ),
                default=0,
            )
            ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, 32)

    # Hoja 1: Sectores
    ws1 = wb_out.active
    ws1.title = "Sectores"
    for i, h in enumerate(
        ["Sector", "Equipo", "Nro Sector", "Ha Totales", "Caudal (m3/h)", "Caseta"], 1
    ):
        ws1.cell(row=1, column=i, value=h)
    r1 = conn.execute("SELECT * FROM sectores ORDER BY sector_nom").fetchall()
    for r, row in enumerate(r1, 2):
        for c, k in enumerate(
            [
                "sector_nom",
                "equipo",
                "sector_num",
                "has_total",
                "caudal_nominal",
                "caseta",
            ],
            1,
        ):
            cell = ws1.cell(row=r, column=c, value=row[k])
            cell.border = BDR
            if k == "has_total":
                cell.number_format = "0.00"
    style_h(ws1, 6)
    auto_w(ws1, 6, len(r1) + 1)

    # Hoja 2: Cuartel x Sector
    ws2 = wb_out.create_sheet("Cuartel_x_Sector")
    for i, h in enumerate(
        [
            "CC",
            "Variedad",
            "Anio",
            "Dist. Hilera",
            "Dist. Plantas",
            "Ha Cuartel",
            "Sector",
            "% del Cuartel",
            "Ha en Sector",
        ],
        1,
    ):
        ws2.cell(row=1, column=i, value=h)
    r2 = conn.execute("""SELECT cs.cuartel_id, c.variedad, c.anio, c.dist_hilera, c.dist_plantas, c.has_total, cs.sector_nom, cs.porcentaje, cs.has_en_sector
                          FROM cuartel_sector cs JOIN cuarteles c ON cs.cuartel_id = c.cc ORDER BY cs.cuartel_id, cs.sector_nom""").fetchall()
    for r, row in enumerate(r2, 2):
        for c, v in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = BDR
            cell.alignment = BA
            if c in (6, 9):
                cell.number_format = "0.00"
            if c == 8:
                cell.number_format = "0%"
        if r % 2 == 0:
            for c in range(1, 10):
                ws2.cell(row=r, column=c).fill = ALT
    style_h(ws2, 9)
    auto_w(ws2, 9, len(r2) + 1)

    # Hoja 3: Historial Riegos
    ws3 = wb_out.create_sheet("Historial_Riegos")
    for i, h in enumerate(
        [
            "ID",
            "Fecha",
            "Sector",
            "Volumen (m3)",
            "Duracion (h)",
            "Hora Ini",
            "Hora Fin",
            "Archivo",
        ],
        1,
    ):
        ws3.cell(row=1, column=i, value=h)
    r3 = conn.execute(
        "SELECT id, fecha, sector_nom, volumen_m3, duracion_h, hora_ini, hora_fin, archivo_origen FROM riegos ORDER BY fecha"
    ).fetchall()
    for r, row in enumerate(r3, 2):
        for c, v in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = BDR
            if c in (4, 5):
                cell.number_format = "#,##0.0"
        if r % 2 == 0:
            for c in range(1, 9):
                ws3.cell(row=r, column=c).fill = ALT
    style_h(ws3, 8)
    auto_w(ws3, 8, len(r3) + 1)

    meses = conn.execute(
        "SELECT DISTINCT strftime('%Y-%m', fecha) as mes FROM riegos ORDER BY mes"
    ).fetchall()
    mes_raw = [m["mes"] for m in meses] if meses else []
    mes_disp = [f"{MESES_ES[ml.split('-')[1]]} {ml.split('-')[0]}" for ml in mes_raw]

    cuarteles = conn.execute("""SELECT rc.cuartel_id, c.variedad, cs.has_en_sector as ha_regada,
        rc.sector_nom, cs.equipo,
        (rc.cuartel_id || '-' || rc.sector_nom) as id_unico
        FROM riegos_cuartel rc
        JOIN cuarteles c ON rc.cuartel_id = c.cc
        JOIN cuartel_sector cs ON rc.cuartel_id = cs.cuartel_id AND rc.sector_nom = cs.sector_nom
        GROUP BY rc.cuartel_id, rc.sector_nom
        ORDER BY rc.cuartel_id, rc.sector_nom""").fetchall()

    m3_dict = {}
    for row in (
        conn.execute("""SELECT rc.cuartel_id, rc.sector_nom, strftime('%Y-%m', r.fecha) as mes,
                               ROUND(SUM(rc.volumen_m3),1) as m3
                               FROM riegos_cuartel rc JOIN riegos r ON rc.riego_id = r.id
                               GROUP BY rc.cuartel_id, rc.sector_nom, mes""").fetchall()
    ):
        m3_dict[(row["cuartel_id"], row["sector_nom"], row["mes"])] = row["m3"]

    # Hoja 4: Resumen Simple (normalizado)
    ws4 = wb_out.create_sheet("Resumen_Cuartel_x_Mes")
    MI = 7  # columna donde empiezan los meses
    h4 = (
        ["ID", "CC", "Variedad", "Ha en Sector", "Equipo", "Sector"]
        + [f"{m} (m3)" for m in mes_disp]
        + ["Total (m3)", "m3/ha"]
    )
    nc4 = len(h4)
    for i, h in enumerate(h4, 1):
        ws4.cell(row=1, column=i, value=h)
    for r, cd in enumerate(cuarteles, 2):
        cc, has, sn, eqn, uid = (
            cd["cuartel_id"],
            cd["ha_regada"],
            cd["sector_nom"],
            cd["equipo"],
            cd["id_unico"],
        )
        for c, v in enumerate([uid, cc, cd["variedad"], has, eqn, sn], 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.border = BDR
            if c == 4:
                cell.number_format = "0.00"
        total_cc = 0
        for ci, mk in enumerate(mes_raw):
            m3 = m3_dict.get((cc, sn, mk), 0)
            col = MI + ci
            cell = ws4.cell(row=r, column=col, value=m3 if m3 else None)
            cell.border = BDR
            cell.alignment = BA
            cell.number_format = "#,##0.0"
            if m3:
                total_cc += m3
        ct = ws4.cell(row=r, column=MI + len(mes_raw), value=round(total_cc, 1))
        ct.border = BDR
        ct.font = Font(bold=True)
        ct.number_format = "#,##0.0"
        ch = ws4.cell(
            row=r,
            column=MI + len(mes_raw) + 1,
            value=round(total_cc / has, 1) if has else None,
        )
        ch.border = BDR
        ch.font = Font(bold=True)
        ch.number_format = "#,##0.0"
        if r % 2 == 0:
            for c in range(1, nc4 + 1):
                ws4.cell(row=r, column=c).fill = ALT
    tr = len(cuarteles) + 2
    for c in range(1, nc4 + 1):
        ws4.cell(row=tr, column=c).fill = HFL
        ws4.cell(row=tr, column=c).border = BDR
        ws4.cell(row=tr, column=c).font = Font(bold=True, color="FFFFFF")
    ws4.cell(row=tr, column=1, value="TOTAL")
    for ci, mk in enumerate(mes_raw):
        s = sum(
            m3_dict.get((cd["cuartel_id"], cd["sector_nom"], mk), 0) for cd in cuarteles
        )
        cell = ws4.cell(row=tr, column=MI + ci, value=round(s, 1))
        cell.number_format = "#,##0.0"
    style_h(ws4, nc4)
    auto_w(ws4, nc4, tr + 1)
    ws4.auto_filter.ref = f"A1:{get_column_letter(nc4)}{tr - 1}"

    # Hoja 5: Resumen Detallado (normalizado)
    ws5 = wb_out.create_sheet("Resumen_Detallado")
    MI5 = 7
    h5h = ["ID", "CC", "Variedad", "Ha en Sector", "Equipo", "Sector"]
    for m in mes_disp:
        h5h += [f"{m} (m3)", f"{m} (m3/ha)"]
    h5h += ["Total (m3)", "m3/ha temp"]
    nc5 = len(h5h)
    for i, h in enumerate(h5h, 1):
        ws5.cell(row=1, column=i, value=h)
    for r, cd in enumerate(cuarteles, 2):
        cc, has, sn, eqn, uid = (
            cd["cuartel_id"],
            cd["ha_regada"],
            cd["sector_nom"],
            cd["equipo"],
            cd["id_unico"],
        )
        for c, v in enumerate([uid, cc, cd["variedad"], has, eqn, sn], 1):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.border = BDR
            if c == 4:
                cell.number_format = "0.00"
        total_cc = 0
        for ci, mk in enumerate(mes_raw):
            cm3 = MI5 + ci * 2
            cmha = cm3 + 1
            m3 = m3_dict.get((cc, sn, mk), 0)
            c1 = ws5.cell(row=r, column=cm3, value=m3 if m3 else None)
            c1.border = BDR
            c1.alignment = BA
            c1.number_format = "#,##0.0"
            mha = round(m3 / has, 1) if has and m3 else None
            c2 = ws5.cell(row=r, column=cmha, value=mha)
            c2.border = BDR
            c2.alignment = BA
            c2.number_format = "#,##0.0"
            c2.font = Font(bold=True, size=12)
            if m3:
                total_cc += m3
        ct = ws5.cell(row=r, column=MI5 + len(mes_raw) * 2, value=round(total_cc, 1))
        ct.border = BDR
        ct.font = Font(bold=True)
        ct.number_format = "#,##0.0"
        ch = ws5.cell(
            row=r,
            column=MI5 + len(mes_raw) * 2 + 1,
            value=round(total_cc / has, 1) if has else None,
        )
        ch.border = BDR
        ch.font = Font(bold=True)
        ch.number_format = "#,##0.0"
        if r % 2 == 0:
            for c in range(1, nc5 + 1):
                ws5.cell(row=r, column=c).fill = ALT
    tr5 = len(cuarteles) + 2
    for c in range(1, nc5 + 1):
        ws5.cell(row=tr5, column=c).fill = HFL
        ws5.cell(row=tr5, column=c).border = BDR
        ws5.cell(row=tr5, column=c).font = Font(bold=True, color="FFFFFF")
    ws5.cell(row=tr5, column=1, value="TOTAL")
    for ci, mk in enumerate(mes_raw):
        cm3 = MI5 + ci * 2
        s = sum(
            m3_dict.get((cd["cuartel_id"], cd["sector_nom"], mk), 0) for cd in cuarteles
        )
        cell = ws5.cell(row=tr5, column=cm3, value=round(s, 1))
        cell.number_format = "#,##0.0"
    style_h(ws5, nc5)
    auto_w(ws5, nc5, tr5 + 1)

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    n_riegos = conn.execute("SELECT COUNT(*) FROM riegos").fetchone()[0]
    n_cr = conn.execute("SELECT COUNT(*) FROM riegos_cuartel").fetchone()[0]
    vol = conn.execute("SELECT COALESCE(SUM(volumen_m3),0) FROM riegos").fetchone()[0]
    conn.close()
    return output, n_riegos, n_cr, vol


# ─── UI ───────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(
    ["1. Limpiar Datos", "2. Auditar Riegos", "3. Configurar Cuarteles"]
)

# ─── TAB 1: LIMPIAR DATOS ─────────────────────────────
with tab1:
    st.subheader("Subi el archivo .xls de Agronic")
    raw_file = st.file_uploader("Archivo crudo (.xls)", type=["xls"], key="raw")
    if raw_file:
        with st.spinner("Limpiando..."):
            df_limpio = clean_agronic(raw_file)
        st.success(
            f"Listo: {len(df_limpio)} registros, {df_limpio['Volumen_m3'].sum():,.1f} m3"
        )
        st.dataframe(df_limpio, use_container_width=True, height=300)
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df_limpio.to_excel(w, sheet_name="Riegos", index=False)
        st.download_button(
            "Descargar Planilla Limpiada",
            buf.getvalue(),
            f"Planilla_Limpiada_{raw_file.name.replace('.xls', '')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ─── TAB 2: AUDITAR RIEGOS ────────────────────────────
with tab2:
    sectores_df, cuartel_sector_df = load_sectores()
    cols = st.columns([1, 1, 1])
    equipo = cols[0].number_input("Equipo (0 = TODOS)", 0, 30, 1)
    cols[1].metric(
        "Sectores configurados",
        len(
            sectores_df[sectores_df["equipo"] == equipo] if equipo > 0 else sectores_df
        ),
    )
    cols[2].metric(
        "CC-Sector activos",
        len(
            cuartel_sector_df[
                (cuartel_sector_df["equipo"] == equipo if equipo > 0 else True)
                & cuartel_sector_df["activo"]
            ]
        ),
    )

    st.subheader("Subi las planillas de riego")
    riegos_files = st.file_uploader(
        "Planillas .xlsx o .xls",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="riegos2",
    )

    if riegos_files and st.button("Procesar Auditoria", type="primary"):
        cleaned = []
        for f in riegos_files:
            if f.name.endswith(".xls"):
                with st.spinner(f"Limpiando {f.name}..."):
                    df_clean = clean_agronic(f)
                    buf = BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as w:
                        df_clean.to_excel(w, sheet_name="Riegos", index=False)
                    buf.seek(0)
                    buf.name = f.name.replace(".xls", ".xlsx")
                    cleaned.append(buf)
            else:
                cleaned.append(f)

        with st.spinner("Ejecutando auditoria..."):
            out, n_riegos, n_cr, vol = run_audit(
                sectores_df, cuartel_sector_df, cleaned, equipo if equipo > 0 else None
            )

        st.success(
            f"Auditoria: {n_riegos} riegos, {n_cr} registros cuartel, {vol:,.1f} m3"
        )
        eq_label = f"Equipo{equipo}" if equipo > 0 else "Completo"
        st.download_button(
            f"Descargar Auditoria_Riego_{eq_label}.xlsx",
            out.getvalue(),
            f"Auditoria_Riego_{eq_label}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ─── TAB 3: CONFIGURAR CUARTELES ──────────────────────
with tab3:
    st.subheader("Base de datos interna de Cuarteles y Sectores")
    st.caption(
        "Acá podés ver y editar la relacion cuartel-sector. Los cambios se usan en la auditoria."
    )

    sectores_df, cuartel_sector_df = load_sectores()

    sub1, sub2 = st.tabs(["Cuartel x Sector", "Sectores"])

    with sub1:
        st.caption(f"{len(cuartel_sector_df)} relaciones cuartel-sector")
        activos_count = (
            int(cuartel_sector_df["activo"].sum())
            if "activo" in cuartel_sector_df.columns
            else len(cuartel_sector_df)
        )
        st.caption(
            f"{activos_count} activos / {len(cuartel_sector_df) - activos_count} inactivos"
        )

        # Filtros
        with st.expander("Filtros", expanded=False):
            f1, f2, f3, f4 = st.columns(4)
            equipos_opts = sorted(
                cuartel_sector_df["equipo"].dropna().unique().astype(int)
            )
            eq_filter = f1.multiselect("Equipo", equipos_opts, placeholder="Todos")

            sectores_opts = sorted(
                cuartel_sector_df["sector"].dropna().unique().astype(int)
            )
            sec_filter = f2.multiselect("Sector", sectores_opts, placeholder="Todos")

            cc_opts = sorted(cuartel_sector_df["cc"].dropna().unique().astype(int))
            cc_filter = f3.multiselect("Cuartel (CC)", cc_opts, placeholder="Todos")

            var_opts = sorted(cuartel_sector_df["variedad"].dropna().unique())
            var_filter = f4.multiselect("Variedad", var_opts, placeholder="Todas")

            activo_filter = st.checkbox("Mostrar solo activos", value=False)

        df_mostrar = cuartel_sector_df.copy()
        if eq_filter:
            df_mostrar = df_mostrar[df_mostrar["equipo"].astype(int).isin(eq_filter)]
        if sec_filter:
            df_mostrar = df_mostrar[df_mostrar["sector"].astype(int).isin(sec_filter)]
        if cc_filter:
            df_mostrar = df_mostrar[df_mostrar["cc"].astype(int).isin(cc_filter)]
        if var_filter:
            df_mostrar = df_mostrar[df_mostrar["variedad"].isin(var_filter)]
        if activo_filter:
            df_mostrar = df_mostrar[df_mostrar["activo"] == True]

        st.caption(f"Mostrando {len(df_mostrar)} registros")

        edited_cs = st.data_editor(
            df_mostrar,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "id": st.column_config.TextColumn("ID", disabled=True),
                "cc": st.column_config.NumberColumn("CC", min_value=1, step=1),
                "variedad": st.column_config.TextColumn("Variedad"),
                "dh": st.column_config.NumberColumn("Dist. Hilera", format="%.1f"),
                "dp": st.column_config.NumberColumn("Dist. Plantas", format="%.1f"),
                "anio": st.column_config.NumberColumn("Anio", min_value=2000, step=1),
                "equipo": st.column_config.NumberColumn("Equipo", min_value=1, step=1),
                "sector": st.column_config.NumberColumn("Sector", min_value=1, step=1),
                "pct": st.column_config.NumberColumn(
                    "% en Sector", format="%.2f", min_value=0.0, max_value=1.0
                ),
                "has_total": st.column_config.NumberColumn("Ha Cuartel", format="%.2f"),
                "has_en_sector": st.column_config.NumberColumn(
                    "Ha en Sector", format="%.2f"
                ),
                "activo": st.column_config.CheckboxColumn("Activo", default=True),
            },
            column_order=[
                "id",
                "cc",
                "variedad",
                "equipo",
                "sector",
                "pct",
                "has_total",
                "has_en_sector",
                "dh",
                "dp",
                "anio",
                "activo",
            ],
            key="edit_cuartel_sector",
        )
        if not edited_cs.equals(df_mostrar):
            # Push current state to undo stack (max 4)
            if "undo_stack" not in st.session_state:
                st.session_state.undo_stack = []
            current = cuartel_sector_df.copy()
            st.session_state.undo_stack.append(current)
            if len(st.session_state.undo_stack) > 4:
                st.session_state.undo_stack.pop(0)

            # Merge edits back into full DF
            full_df = cuartel_sector_df.copy()
            for idx in edited_cs.index:
                full_df.loc[idx] = edited_cs.loc[idx]
            st.session_state.cuartel_sector_df = full_df
            st.success("Cuartel x Sector actualizado")
            st.rerun()

        # Undo button
        undo_avail = len(st.session_state.get("undo_stack", []))
        col_undo, col_empty = st.columns([1, 5])
        if col_undo.button(
            f"Deshacer ({undo_avail})",
            disabled=undo_avail == 0,
            help="Deshacer el ultimo cambio (hasta 4 veces)",
        ):
            prev = st.session_state.undo_stack.pop()
            st.session_state.cuartel_sector_df = prev
            st.success("Cambio deshecho")
            st.rerun()

    with sub2:
        st.caption(f"{len(sectores_df)} sectores configurados")
        edited_sectores = st.data_editor(
            sectores_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "equipo": st.column_config.NumberColumn(
                    "Equipo", min_value=1, max_value=50, step=1
                ),
                "sector": st.column_config.NumberColumn("Sector", min_value=1, step=1),
                "caudal": st.column_config.NumberColumn("Caudal (m3/h)", format="%.1f"),
                "has": st.column_config.NumberColumn("Ha", format="%.2f"),
                "caseta": st.column_config.TextColumn("Caseta"),
            },
            key="edit_sectores",
        )
        if not edited_sectores.equals(sectores_df):
            st.session_state.sectores_df = edited_sectores
            st.success("Sectores actualizados")

    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "Descargar configuracion actual (Excel)",
            save_session_data().getvalue(),
            "Cuartel_x_Sector_Config.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col2:
        uploaded_config = st.file_uploader(
            "Cargar configuracion desde Excel", type=["xlsx"], key="upload_config"
        )
        if uploaded_config:
            wb_up = openpyxl.load_workbook(uploaded_config, data_only=True)
            if (
                "Sectores" in wb_up.sheetnames
                and "Cuartel x Sector" in wb_up.sheetnames
            ):
                ws_s = wb_up["Sectores"]
                new_s = []
                for row in ws_s.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    new_s.append(
                        {
                            "equipo": int(row[0]),
                            "sector": int(row[1]),
                            "caudal": row[2],
                            "has": row[3],
                            "caseta": str(row[4]) if row[4] else None,
                        }
                    )
                ws_cs = wb_up["Cuartel x Sector"]
                # Detectar formato: col 0 = "ID" (nuevo) o "CC" (viejo)
                first_header = str(ws_cs.cell(row=1, column=1).value or "").strip()
                has_id_col = first_header == "ID"
                new_cs = []
                for row in ws_cs.iter_rows(min_row=2, values_only=True):
                    ofs = 1 if has_id_col else 0  # offset por columna ID
                    if row[ofs] is None or row[5 + ofs] is None:
                        continue
                    cc = int(row[ofs])
                    eq = int(row[5 + ofs])
                    sec = int(row[6 + ofs])
                    sn = f"E{eq}S{sec}"
                    new_cs.append(
                        {
                            "id": str(row[0]).strip()
                            if has_id_col and row[0]
                            else f"{cc}-{sn}",
                            "cc": cc,
                            "variedad": str(row[1 + ofs]),
                            "dh": row[2 + ofs],
                            "dp": row[3 + ofs],
                            "anio": int(row[4 + ofs]) if row[4 + ofs] else None,
                            "equipo": int(row[5 + ofs]),
                            "sector": int(row[6 + ofs]),
                            "pct": row[7 + ofs],
                            "has_total": row[8 + ofs],
                            "has_en_sector": row[9 + ofs],
                            "activo": bool(row[10 + ofs])
                            if len(row) > 10 + ofs and row[10 + ofs] is not None
                            else True,
                        }
                    )
                st.session_state.sectores_df = pd.DataFrame(new_s)
                st.session_state.cuartel_sector_df = pd.DataFrame(new_cs)
                st.success(
                    f"Configuracion cargada: {len(new_s)} sectores, {len(new_cs)} relaciones"
                )
                st.rerun()
            else:
                st.error("El Excel debe tener hojas 'Sectores' y 'Cuartel x Sector'")
